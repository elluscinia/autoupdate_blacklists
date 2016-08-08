import unittest
import win32com.client
from uuid import uuid4
from xml.sax.saxutils import unescape
import time
import sys
import os
import openpyxl
import pandas as pd
from openpyxl.chart import Reference, Series, ScatterChart

xmlStart = """<?xml version="1.0" encoding="UTF-8"?><fpc4:Root xmlns:fpc4="http://schemas.microsoft.com/isa/config-4" xmlns:dt="urn:schemas-microsoft-com:datatypes" StorageName="FPC" StorageType="0"><fpc4:Build dt:dt="string">7.0.7734.100</fpc4:Build><fpc4:Comment dt:dt="string"/><fpc4:Edition dt:dt="int">32</fpc4:Edition><fpc4:EnterpriseLevel dt:dt="int">2</fpc4:EnterpriseLevel><fpc4:ExportItemClassCLSID dt:dt="string">{61A8568E-53C1-4D6D-BBD8-4F7150EB3093}</fpc4:ExportItemClassCLSID><fpc4:ExportItemCompatibilityVersion dt:dt="int">2</fpc4:ExportItemCompatibilityVersion><fpc4:ExportItemScope dt:dt="int">0</fpc4:ExportItemScope><fpc4:ExportItemStorageName dt:dt="string">{%(expGUID)s}</fpc4:ExportItemStorageName><fpc4:IsaXmlVersion dt:dt="string">7.3</fpc4:IsaXmlVersion><fpc4:OptionalData dt:dt="int">12</fpc4:OptionalData><fpc4:Upgrade dt:dt="boolean">0</fpc4:Upgrade><fpc4:ConfigurationMode dt:dt="int">0</fpc4:ConfigurationMode><fpc4:Arrays StorageName="Arrays" StorageType="0"><fpc4:Array StorageName="{9DABC2DD-2B86-4200-B856-F755E7441696}" StorageType="0"><fpc4:AdminMajorVersion dt:dt="int">0</fpc4:AdminMajorVersion><fpc4:AdminMinorVersion dt:dt="int">0</fpc4:AdminMinorVersion><fpc4:Components dt:dt="int">-1</fpc4:Components><fpc4:DNSName dt:dt="string"/><fpc4:Name dt:dt="string"/><fpc4:Version dt:dt="string">0</fpc4:Version><fpc4:RuleElements StorageName="RuleElements" StorageType="0"><fpc4:DomainNameSets StorageName="DomainNameSets" StorageType="0"><fpc4:DomainNameSet StorageName="{%(expGUID)s}" StorageType="1"><fpc4:DomainNameStrings>"""

xmlEnd = """</fpc4:DomainNameStrings><fpc4:Name dt:dt="string">%s</fpc4:Name></fpc4:DomainNameSet></fpc4:DomainNameSets></fpc4:RuleElements></fpc4:Array></fpc4:Arrays></fpc4:Root>"""


def parse_statistics(logfile):
    xl = pd.ExcelFile(logfile)
    df = xl.parse("Sheet")
    df = df.sort_values(by='Line Numbers')

    writer = pd.ExcelWriter(logfile)
    df.to_excel(writer, sheet_name='Sheet', index=False)
    writer.save()

    wb = openpyxl.load_workbook(logfile)
    ws = wb.active

    row_count = ws.max_row
    column_count = ws.max_column

    chart = ScatterChart()
    chart.title = "Time upload domain names"
    chart.style = 13
    chart.x_axis.title = "Line numbers"
    chart.y_axis.title = "Time, sec"

    xvalues = Reference(ws, min_col=1, min_row=2, max_row=row_count)
    color_choice = ['3F888F', 'D24D57']
    for i in range(2, column_count + 1):
        values = Reference(ws, min_col=i, min_row=1, max_row=row_count)
        series = Series(values, xvalues, title_from_data=True)
        series.marker.symbol = "diamond"
        series.graphicalProperties.line.solidFill = color_choice[i-2]
        series.marker.graphicalProperties.line.solidFill = color_choice[i-2]
        series.marker.graphicalProperties.solidFill = color_choice[i-2]
        series.graphicalProperties.line.width = 20000
        chart.series.append(series)

    chart.legend.legendPos = 'b'
    ws.add_chart(chart)
    wb.save(logfile)


class TestCompareMethods(unittest.TestCase):
    def __init__(self, testname, rulename, filename, logfile):
        super(TestCompareMethods, self).__init__(testname)
        self.rule_name = rulename
        self.file_name = filename
        self.logfile = logfile

    def upload_xml(self):
        start_time = time.time()
        f = open(self.file_name)
        domains = [domain for domain in f]
        f.close()
        domain_name = self.file_name + '_xml'
        rule_set = self.rule_name + '_XML'

        file_dom = (xmlStart % {'expGUID': str(uuid4()).upper()})
        for url in domains:
            file_dom += ('<fpc4:Str dt:dt="string">' + (url.replace('&', '&amp;')) + '</fpc4:Str>')
        file_dom += (xmlEnd % unescape(domain_name))

        dom = win32com.client.Dispatch('Msxml2.DOMDocument.3.0')
        dom.async = False
        dom.loadXML(file_dom)

        isa_array.RuleElements.DomainNameSets.Import(dom, 0)

        rule = isa_array.ArrayPolicy.PolicyRules.Item(rule_set)

        rule_sets = rule.AccessProperties.DestinationDomainNameSets

        rule_sets.Add(domain_name, 0)

        rule.Save()

        return time.time() - start_time

    def upload_each_domain(self):
        start_time = time.time()
        f = open(self.file_name)
        domains = [domain for domain in f]
        f.close()
        domain_name = self.file_name + '_each_domain'
        rule_set = self.rule_name + '_EachDomain'

        domain_name_set = isa_array.RuleElements.DomainNameSets.Add(domain_name)
        domain_name_set.Save()
        domain_name_set = isa_array.RuleElements.DomainNameSets.Item(domain_name)

        for url in domains:
            try:
                domain_name_set.Add(url)
            except Exception as ex:
                pass

        domain_name_set.Save()

        rule = isa_array.ArrayPolicy.PolicyRules.Item(rule_set)

        rule_sets = rule.AccessProperties.DestinationDomainNameSets

        rule_sets.Add(domain_name, 0)

        rule.Save()
        return time.time() - start_time

    def test_compare(self):
        wb = openpyxl.load_workbook(self.logfile)
        ws = wb.active

        file_strings = sum(1 for l in open(self.file_name, 'r'))
        time_upload_each_domain = self.upload_each_domain()
        time_upload_xml = self.upload_xml()

        ws.append([file_strings, time_upload_each_domain, time_upload_xml])

        wb.save(self.logfile)

        self.assertGreater(time_upload_each_domain, time_upload_xml)

        object_tmg = win32com.client.Dispatch('FPC.Root')
        isa_array = object_tmg.GetContainingArray()
        i = 1
        while (i <= isa_array.ArrayPolicy.PolicyRules.Count):
            rule = isa_array.ArrayPolicy.PolicyRules.Item(i).Name
            if rule.startswith(self.rule_name):
                isa_array.ArrayPolicy.PolicyRules.Remove(i)
            else:
                i += 1
        isa_array.ArrayPolicy.PolicyRules.Save()
        i = 1
        while (i <= isa_array.RuleElements.DomainNameSets.Count):
            domain_set = isa_array.RuleElements.DomainNameSets.Item(i).Name
            if domain_set.startswith(self.file_name):
                isa_array.RuleElements.DomainNameSets.Remove(i)
            else:
                i += 1
        isa_array.RuleElements.DomainNameSets.Save()

if __name__ == '__main__':

    directory = sys.argv[1]
    rulename = 'TestCase'
    log_file = 'statistics.xlsx'

    headers = ['Line Numbers', 'Time upload each domain, sec', 'Time XML import, sec',]

    wb = openpyxl.workbook.Workbook()
    ws1 = wb.active
    ws1.append(headers)

    wb.save(filename=log_file)

    object_tmg = win32com.client.Dispatch('FPC.Root')
    isa_array = object_tmg.GetContainingArray()
    isa_array.ArrayPolicy.PolicyRules.AddAccessRule(rulename + '_EachDomain')
    isa_array.ArrayPolicy.PolicyRules.AddAccessRule(rulename + '_XML')
    isa_array.ArrayPolicy.PolicyRules.Save()

    suite = unittest.TestSuite()
    for root, dirs, files in os.walk(directory):
        for filename in files:
            suite.addTest(TestCompareMethods('test_compare', rulename, os.path.join(root, filename), log_file))
    unittest.TextTestRunner().run(suite)

    parse_statistics(log_file)

